from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
# from flask_login import login_user, logout_user, login_required, current_user  # 已注释掉登录相关导入
from datetime import datetime, timedelta
from app.models import User, QueryLog
from app.api_client import APIClient
from io import BytesIO
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 创建蓝图
main_bp = Blueprint('main', __name__)
auth_bp = Blueprint('auth', __name__)

# 认证相关路由 - 已注释掉登录功能
# @auth_bp.route('/login', methods=['GET', 'POST'])
# def login():
#     # 调试信息
#     print(f"Login request - URL: {request.url}")
#     print(f"Login request - Headers: {dict(request.headers)}")
#     
#     if request.method == 'POST':
#         username = request.form['username']
#         password = request.form['password']
#         
#         user = User.get_by_username(username)
#         if user and user.check_password(password):
#             login_user(user)
#             next_page = request.args.get('next')
#             
#             # 调试重定向
#             if next_page:
#                 print(f"Redirecting to next_page: {next_page}")
#                 return redirect(next_page)
#             else:
#                 dashboard_url = url_for('main.dashboard')
#                 print(f"Redirecting to dashboard: {dashboard_url}")
#                 return redirect(dashboard_url)
#         else:
#             flash('用户名或密码错误', 'error')
#     
#     return render_template('login.html')

# @auth_bp.route('/logout')
# @login_required
# def logout():
#     logout_user()
#     flash('您已成功退出登录', 'success')
#     return redirect(url_for('auth.login'))

# 主页面路由 - 已移除登录要求
@main_bp.route('/')
# @login_required  # 已注释掉登录要求
def dashboard():
    # 获取用户最近的查询记录 - 暂时注释掉用户相关功能
    # recent_logs = QueryLog.get_user_logs(current_user.id, limit=5)
    recent_logs = []  # 临时设置为空列表
    return render_template('dashboard.html', recent_logs=recent_logs)

@main_bp.route('/query', methods=['GET', 'POST'])
# @login_required  # 已注释掉登录要求
def query_data():
    if request.method == 'POST':
        date_input = request.form.get('date')
        if not date_input:
            flash('请输入查询日期', 'error')
            return render_template('query.html')
        
        # 格式化日期
        api_client = APIClient()
        formatted_date = api_client.format_date(date_input)
        
        # 调用API获取数据
        raw_data = api_client.get_attribution_data(formatted_date)
        if raw_data is None:
            flash('获取数据失败，请检查日期格式或网络连接', 'error')
            return render_template('query.html')
        
        # 解析数据
        parsed_data = api_client.parse_api_response(raw_data)
        if parsed_data is None:
            flash('数据解析失败', 'error')
            return render_template('query.html')
        
        # 保存查询记录到数据库 - 暂时注释掉用户相关功能
        # summary = parsed_data.get('summary', {})
        # QueryLog.create(
        #     current_user.id,
        #     formatted_date,
        #     summary.get('total_request_count', 0),
        #     summary.get('total_error_count', 0)
        # )
        
        return render_template('query_result.html', data=parsed_data, query_date=formatted_date)
    
    return render_template('query.html')

@main_bp.route('/api/data')
# @login_required  # 已注释掉登录要求
def api_data():
    """提供JSON格式的数据接口"""
    date = request.args.get('date')
    if not date:
        return jsonify({'error': '缺少日期参数'}), 400
    
    api_client = APIClient()
    formatted_date = api_client.format_date(date)
    
    raw_data = api_client.get_attribution_data(formatted_date)
    if raw_data is None:
        return jsonify({'error': '获取数据失败'}), 500
    
    parsed_data = api_client.parse_api_response(raw_data)
    if parsed_data is None:
        return jsonify({'error': '数据解析失败'}), 500
    
    return jsonify(parsed_data)


@main_bp.route('/profile')
# @login_required  # 已注释掉登录要求
def profile():
    """用户资料页面"""
    return render_template('profile.html')

@main_bp.route('/export/error-counts-excel')
# @login_required  # 已注释掉登录要求
def export_error_counts():
    """导出过去10天的错误统计数据为Excel"""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': '缺少openpyxl库，请先安装'}), 500
    
    try:
        api_client = APIClient()
        
        # 获取过去10天的数据
        multi_days_data = api_client.get_multi_days_error_data(num_days=10)
        
        if not multi_days_data:
            return jsonify({'error': '无法获取数据，请稍后重试'}), 500
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "错误统计"
        
        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        # 写入标题
        headers = ["日期", "账户ID", "错误类型", "事件", "错误数量"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # 按日期降序排列（从最新开始）
        sorted_dates = sorted(multi_days_data.keys(), reverse=True)
        
        row = 2
        for date_str in sorted_dates:
            error_counts = multi_days_data[date_str]
            
            # 按账户ID排序
            for advertiser_id in sorted(error_counts.keys()):
                errors = error_counts[advertiser_id]
                
                # 按错误类型排序
                for error_type in sorted(errors.keys()):
                    error_count = errors[error_type]
                    
                    # 拆分错误类型和事件
                    # 例如: "advertiser_rate_false_1" -> ("advertiser_rate_false", "1")
                    # 如果没有下划线+数字后缀，事件列为空
                    parts = error_type.rsplit('_', 1)
                    if len(parts) == 2 and parts[1].isdigit():
                        error_category = parts[0]
                        event_number = parts[1]
                    else:
                        error_category = error_type
                        event_number = ""
                    
                    # 写入数据
                    ws.cell(row=row, column=1).value = date_str
                    ws.cell(row=row, column=2).value = advertiser_id
                    ws.cell(row=row, column=3).value = error_category
                    ws.cell(row=row, column=4).value = event_number
                    ws.cell(row=row, column=5).value = error_count
                    
                    # 应用样式
                    for col in range(1, 6):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        if col in [4, 5]:  # 数字列居中
                            cell.alignment = center_alignment
                    
                    row += 1
        
        # 在内存中创建 Excel 文件
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 生成文件名
        export_date = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"error_counts_{export_date}.xlsx"
        
        # 返回 Excel 文件
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"导出Excel出错: {e}")
        return jsonify({'error': f'导出失败: {str(e)}'}), 500

